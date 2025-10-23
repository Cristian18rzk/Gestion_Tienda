# persistencia.py
import csv
import json
from builtins import FileNotFoundError
from typing import List, Dict

# Para exportar Excel/PDF
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer


# =======================
# Lógica de Persistencia CSV
# =======================

class PersistenciaCSV:
    """Maneja la lectura y escritura en archivos CSV para Productos y Clientes."""

    @staticmethod
    def leer_datos(nombre_archivo, campos):
        datos = []
        try:
            with open(nombre_archivo, mode='r', newline='', encoding='utf-8') as file:
                # Usamos DictReader para leer filas como diccionarios
                reader = csv.DictReader(file, fieldnames=campos)
                next(reader, None)  # Saltar la línea de encabezado si existe
                for row in reader:
                    datos.append(row)
        except FileNotFoundError:
            # Crea el archivo con encabezados si no existe
            PersistenciaCSV.escribir_datos(nombre_archivo, [], campos)
        return datos

    @staticmethod
    def escribir_datos(nombre_archivo, lista_objetos, campos):
        """Escribe una lista de objetos (con método .to_dict()) al CSV."""
        with open(nombre_archivo, mode='w', newline='', encoding='utf-8') as file:
            writer = csv.DictWriter(file, fieldnames=campos)
            writer.writeheader()
            for obj in lista_objetos:
                writer.writerow(obj.to_dict())


# =======================
# Lógica de Persistencia JSON
# =======================

class PersistenciaJSON:
    """Maneja la lectura y escritura en archivos JSON para Pedidos."""

    @staticmethod
    def leer_pedidos(nombre_archivo):
        try:
            with open(nombre_archivo, 'r', encoding='utf-8') as file:
                return json.load(file)
        except FileNotFoundError:
            return []
        except json.JSONDecodeError:
            print(
                f"[bold yellow]Advertencia:[/bold yellow] Archivo '{nombre_archivo}' vacío o corrupto. Inicializando lista de pedidos vacía.")
            return []

    @staticmethod
    def escribir_pedidos(nombre_archivo, pedidos):
        with open(nombre_archivo, 'w', encoding='utf-8') as file:
            # Usamos indent=4 para que el JSON sea legible
            json.dump(pedidos, file, indent=4)

    # --------------------------
    # Export / Utilities
    # --------------------------

    @staticmethod
    def exportar_pedidos_excel(nombre_archivo: str, pedidos: List[Dict]):
        """
        Exporta pedidos a un archivo Excel.
        Crea 2 hojas: 'Pedidos' (resumen por pedido) y 'Items' (cada producto por fila vinculando id_pedido).
        """
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Pedidos"

        # Encabezados resumen pedidos
        encabezados = ["id_pedido", "id_cliente", "nombre_cliente", "fecha_pedido", "total_pedido"]
        ws1.append(encabezados)

        for p in pedidos:
            ws1.append([
                p.get("id_pedido"),
                p.get("id_cliente"),
                p.get("nombre_cliente"),
                p.get("fecha_pedido"),
                p.get("total_pedido"),
            ])

        # Ajustar ancho columnas de forma sencilla
        for i, col in enumerate(encabezados, 1):
            ws1.column_dimensions[get_column_letter(i)].width = max(len(col) + 2, 10)

        # Hoja de items
        ws2 = wb.create_sheet(title="Items")
        encabezados_items = ["id_pedido", "id_producto", "nombre", "cantidad", "precio_unitario", "subtotal"]
        ws2.append(encabezados_items)

        for p in pedidos:
            for it in p.get("items", []):
                ws2.append([
                    p.get("id_pedido"),
                    it.get("id_producto"),
                    it.get("nombre"),
                    it.get("cantidad"),
                    it.get("precio_unitario"),
                    it.get("subtotal"),
                ])
        for i, col in enumerate(encabezados_items, 1):
            ws2.column_dimensions[get_column_letter(i)].width = max(len(col) + 2, 10)

        wb.save(nombre_archivo)

    @staticmethod
    def exportar_pedidos_pdf(nombre_archivo: str, pedidos: List[Dict], titulo: str = "Reporte de Pedidos"):
        """
        Exporta un PDF con un resumen de pedidos y una tabla de items.
        Usa reportlab; la salida es básica pero legible.
        """
        doc = SimpleDocTemplate(nombre_archivo, pagesize=landscape(letter), rightMargin=18, leftMargin=18, topMargin=18, bottomMargin=18)
        styles = getSampleStyleSheet()
        flowables = []

        flowables.append(Paragraph(titulo, styles["Title"]))
        flowables.append(Spacer(1, 8))

        # Tabla resumen de pedidos
        pedidos_encabezado = ["ID", "Fecha", "Cliente", "Total"]
        datos_pedidos = [pedidos_encabezado]
        for p in pedidos:
            datos_pedidos.append([
                str(p.get("id_pedido", "")),
                p.get("fecha_pedido", ""),
                p.get("nombre_cliente", ""),
                f"{p.get('total_pedido', 0):.2f}"
            ])

        t = RLTable(datos_pedidos, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ]))
        flowables.append(t)
        flowables.append(Spacer(1, 12))

        # Items: lista todos los items (puede ser larga)
        items_encabezado = ["Pedido ID", "ID Producto", "Nombre", "Cantidad", "Precio unit.", "Subtotal"]
        datos_items = [items_encabezado]
        for p in pedidos:
            for it in p.get("items", []):
                datos_items.append([
                    str(p.get("id_pedido", "")),
                    str(it.get("id_producto", "")),
                    it.get("nombre", ""),
                    str(it.get("cantidad", "")),
                    f"{it.get('precio_unitario', 0):.2f}",
                    f"{it.get('subtotal', 0):.2f}"
                ])

        if len(datos_items) > 1:
            ti = RLTable(datos_items, repeatRows=1)
            ti.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0ea5a4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]))
            flowables.append(Paragraph("Items por pedido", styles["Heading2"]))
            flowables.append(ti)

        doc.build(flowables)

    @staticmethod
    def filtrar_pedidos_por_fecha(pedidos, desde=None, hasta=None):
        """Filtra los pedidos según un rango de fechas (YYYY-MM-DD)."""
        import datetime

        def parse_fecha(fecha_str):
            try:
                return datetime.datetime.strptime(fecha_str, "%Y-%m-%d").date()
            except Exception:
                return None

        desde_fecha = parse_fecha(desde) if desde else None
        hasta_fecha = parse_fecha(hasta) if hasta else None

        pedidos_filtrados = []
        for pedido in pedidos:
            fecha_pedido = parse_fecha(pedido.get("fecha"))
            if not fecha_pedido:
                continue

            if desde_fecha and fecha_pedido < desde_fecha:
                continue
            if hasta_fecha and fecha_pedido > hasta_fecha:
                continue

            pedidos_filtrados.append(pedido)

        return pedidos_filtrados
